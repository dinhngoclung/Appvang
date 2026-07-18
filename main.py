import flet as ft
import pandas as pd
import os
import sys
import re
import unicodedata
from datetime import datetime
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from collections import defaultdict

# ===================================================================
# HÀM 0: CHUẨN HÓA TIẾNG VIỆT
# ===================================================================
def chuan_hoa(text):
    if pd.isna(text) or text is None:
        return ""
    text = str(text).replace('\n',' ').replace('\r',' ')
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c)!= 'Mn')
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    return text

# ===================================================================
# HÀM XỬ LÝ SỐ
# ===================================================================
def parse_number(s):
    if pd.isna(s):
        return 0
    if isinstance(s, (int, float)):
        return s
    s = str(s).strip().replace(',', '').replace(' ', '')
    try:
        return float(s)
    except:
        return 0

# ========= STYLE =========
thin = Side(style='thin')
bd = Border(left=thin,right=thin,top=thin,bottom=thin)
fh = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
fill_xam = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
fill_trang = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
center = Alignment(horizontal='center', vertical='center')
right_center = Alignment(horizontal='right', vertical='center')
center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_sheet(ws, df_day):
    ws.delete_rows(1, ws.max_row)
    hs = ['Ngày', 'Họ Tên', 'CCCD', 'Địa Chỉ', 'Loại Vàng', 'TL Vàng', 'Giá Vàng', 'Thành Tiền']
    for i,h in enumerate(hs):
        c = ws.cell(row=1, column=i+1, value=h)
        c.font = Font(bold=True)
        c.alignment = center_wrap
        c.border = bd
        c.fill = fh
    ws.row_dimensions[1].height = 26

    cur = 2
    groups=[]
    cg=[]
    last_key = None

    for _, rec in df_day.iterrows():
        key = (rec['Ngay'], str(rec['Ten']).strip(), str(rec['CCCD']).strip())
        if key!= last_key:
            if cg: groups.append(cg)
            cg=[rec]
            last_key = key
        else:
            cg.append(rec)
    if cg: groups.append(cg)

    for g in groups:
        n = len(g)
        st = cur
        tot = 0
        for j in range(n):
            rec = g[j]
            rr = cur + j
            if j == 0:
                ws.cell(row=rr, column=1, value=rec['Ngay']).number_format = 'DD/MM/YYYY'
                ws.cell(row=rr, column=2, value=rec['Ten'])
                ws.cell(row=rr, column=3, value=str(rec['CCCD']).split('.')[0] if pd.notna(rec['CCCD']) else "")
                ws.cell(row=rr, column=4, value=rec['DiaChi'])
            
            ws.cell(row=rr, column=5, value=rec['LoaiVang'])
            ws.cell(row=rr, column=6, value=rec['TLVang']).number_format = '0.00'
            ws.cell(row=rr, column=7, value=rec['GiaVang']).number_format = '#,##0'
            ws.cell(row=rr, column=8, value=rec['ThanhTien']).number_format = '#,##0'
            
            for cc in range(1, 9):
                cell = ws.cell(row=rr, column=cc)
                cell.border = bd
                cell.alignment = right_center if cc in [6, 7, 8] else center
            
            tot += rec['ThanhTien']
        
        if n > 1:
            for cc in [1, 2, 3, 4]:
                ws.merge_cells(start_row=st, start_column=cc, end_row=st+n-1, end_column=cc)
                cell = ws.cell(row=st, column=cc)
                cell.alignment = center
        
        rt = cur + n
        for cc in range(1, 8):
            a = ws.cell(row=rt, column=cc, value="")
            a.border = bd
            a.fill = fill_trang
        a = ws.cell(row=rt, column=8, value=tot)
        a.font = Font(bold=True, color="FF0000")
        a.fill = fill_xam
        a.border = bd
        a.number_format = '#,##0'
        a.alignment = right_center
        cur = rt + 1

    cur += 1
    dt0 = df_day.iloc[0]['Ngay']
    ws.merge_cells(start_row=cur, start_column=4, end_row=cur, end_column=8)
    c = ws.cell(row=cur, column=4, value=f"TỔNG HỢP CHI TIẾT HÀNG BÁN NGÀY {dt0.day:02d}/{dt0.month:02d}/{dt0.year}")
    c.font = Font(bold=True)
    c.fill = fh
    c.alignment = center
    c.border = bd
    for cc in [5, 6, 7, 8]:
        ws.cell(row=cur, column=cc).border = bd
    ws.row_dimensions[cur].height = 26
    cur += 1
    
    titles = ["STT", "Loại Vàng", "TL Vàng", "Giá Vàng", "Thành Tiền"]
    for i, t in enumerate(titles):
        a = ws.cell(row=cur, column=4+i, value=t)
        a.font = Font(bold=True)
        a.fill = fh
        a.alignment = center
        a.border = bd
    ws.row_dimensions[cur].height = 26
    cur += 1
    
    agg = defaultdict(lambda: [0, 0, 0])
    for _, rec in df_day.iterrows():
        loai = str(rec['LoaiVang']).strip()
        agg[loai][0] += rec['TLVang']
        agg[loai][1] += rec['ThanhTien']
        if agg[loai][2] == 0:
            agg[loai][2] = rec['GiaVang']
    
    stt = 1
    tong_tt = 0
    for loai, (tl_sum, tt_sum, gia_mau) in agg.items():
        ws.cell(row=cur, column=4, value=stt).border = bd
        ws.cell(row=cur, column=4).alignment = center
        ws.cell(row=cur, column=5, value=loai).border = bd
        ws.cell(row=cur, column=5).alignment = center
        ws.cell(row=cur, column=6, value=round(tl_sum, 2)).border = bd
        ws.cell(row=cur, column=6).number_format = '0.00'
        ws.cell(row=cur, column=6).alignment = right_center
        ws.cell(row=cur, column=7, value=gia_mau).border = bd
        ws.cell(row=cur, column=7).number_format = '#,##0'
        ws.cell(row=cur, column=7).alignment = right_center
        ws.cell(row=cur, column=8, value=tt_sum).border = bd
        ws.cell(row=cur, column=8).number_format = '#,##0'
        ws.cell(row=cur, column=8).alignment = right_center
        tong_tt += tt_sum
        cur += 1
        stt += 1
    
    ws.cell(row=cur, column=4, value="").border = Border(top=thin, bottom=thin, left=thin, right=None)
    ws.cell(row=cur, column=5, value="").border = Border(top=thin, bottom=thin, left=None, right=None)
    ws.cell(row=cur, column=6, value="").border = Border(top=thin, bottom=thin, left=None, right=None)
    ws.cell(row=cur, column=7, value="").border = Border(top=thin, bottom=thin, left=None, right=None)
    a = ws.cell(row=cur, column=8, value=tong_tt)
    a.font = Font(bold=True, color="FF0000")
    a.fill = fill_xam
    a.border = bd
    a.number_format = '#,##0'
    a.alignment = right_center
    
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 15


def process_excel_file(input_path, log_func):
    log_func(f"\n[1] Đọc file: {os.path.basename(input_path)}")
    log_func(f"    Đường dẫn: {input_path}")
    
    df_all = pd.read_excel(input_path, header=None, dtype=str)
    log_func(f" -> OK: File có {df_all.shape[0]} dòng, {df_all.shape[1]} cột")

    header_row = None
    for i in range(min(20, len(df_all))):
        row_text = ' '.join([str(x) for x in df_all.iloc[i, :8] if pd.notna(x)])
        row_norm = chuan_hoa(row_text)
        if any(keyword in row_norm for keyword in ['ngay', 'ten', 'cccd', 'loai vang', 'ky hieu']):
            header_row = i
            break
    
    if header_row is None:
        for i in range(min(20, len(df_all))):
            non_empty = sum([1 for x in df_all.iloc[i, :8] if pd.notna(x) and str(x).strip()!=''])
            if non_empty >= 5:
                header_row = i
                break
    
    if header_row is None:
        header_row = 0
        log_func(" -> Không tìm thấy header rõ ràng, sử dụng dòng đầu tiên")
    else:
        log_func(f" -> Tìm thấy header ở dòng {header_row + 1}")

    df = df_all.iloc[header_row+1:, :8].copy()
    df.columns = ['Ngay', 'Ten', 'CCCD', 'DiaChi', 'LoaiVang', 'TLVang', 'GiaVang', 'ThanhTien']
    
    df = df.dropna(subset=['Ten', 'LoaiVang'], how='all')
    df = df[df['Ten'].astype(str).str.strip() != '']
    
    df['Ngay'] = pd.to_datetime(df['Ngay'], dayfirst=True, errors='coerce')
    if df['Ngay'].isna().sum() > len(df) * 0.5:
        df['Ngay'] = pd.to_datetime(df['Ngay'], errors='coerce')
    
    df = df[df['Ngay'].notna()].copy()
    df = df.sort_values('Ngay').reset_index(drop=True)
    
    log_func(f"[2] Lọc xong: {len(df)} dòng hợp lệ")
    if len(df) > 0:
        log_func(f" -> Ngày đầu: {df.iloc[0]['Ngay'].strftime('%d/%m/%Y')}")
        log_func(f" -> Ngày cuối: {df.iloc[-1]['Ngay'].strftime('%d/%m/%Y')}")

    for col in ['TLVang', 'GiaVang', 'ThanhTien']:
        df[col] = df[col].apply(parse_number)

    df['Ngay_date'] = df['Ngay'].dt.date
    thang, nam = df.iloc[0]['Ngay'].month, df.iloc[0]['Ngay'].year
    
    output_name = f"OKKK-T{thang}-{nam}.xlsx"
    output_dir = os.path.dirname(input_path)
    output_path = os.path.join(output_dir, output_name)
    
    log_func(f"\n[5] Ghi file: {output_path}")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as w:
        df.to_excel(w, sheet_name=f"T{thang}.{nam}", index=False)
        write_sheet(w.sheets[f"T{thang}.{nam}"], df)
        log_func(f"[6] Đã viết sheet tổng: T{thang}.{nam}")
        
        for ng, dfg in df.groupby('Ngay_date'):
            ten = ng.strftime('%d.%m')
            dfg.to_excel(w, sheet_name=ten, index=False)
            write_sheet(w.sheets[ten], dfg)
            log_func(f"[7] Đã viết sheet ngày: {ten}")
    
    log_func(f"\n{'='*40}")
    log_func(f"THÀNH CÔNG!")
    log_func(f"File: {output_path}")
    log_func(f"Size: {os.path.getsize(output_path)/1024:.1f} KB")
    log_func(f"{'='*40}")
    return output_path

# ===================================================================
# GIAO DIỆN FLET - FIX LỖI EOF
# ===================================================================
def main(page: ft.Page):
    page.title = "Tool Tách Sheet - Bản 8 Cột"
    page.scroll = ft.ScrollMode.AUTO
    page.theme_mode = ft.ThemeMode.DARK
    page.padding = 20

    log_text = ft.Text(
        "BẮT ĐẦU CHẠY TOOL TÁCH SHEET - BẢN 8 CỘT FULL GHI CHÚ\nChưa chọn file...\n",
        selectable=True,
        size=13,
        font_family="monospace"
    )
    
    log_container = ft.Container(
        content=ft.Column([log_text], scroll=ft.ScrollMode.AUTO, auto_scroll=True),
        height=400,
        padding=10,
        bgcolor="#1a1a1a",
        border_radius=8,
        border=ft.border.all(1, "#333")
    )

    def append_log(msg):
        log_text.value += "\n" + msg
        page.update()
        # auto scroll xuống cuối
        if log_container.content:
            log_container.content.scroll_to(offset=-1, duration=200)

    def on_pick_result(e: ft.FilePickerResultEvent):
        if not e.files:
            append_log("Bạn chưa chọn file nào.")
            return
        
        file_path = e.files[0].path
        if not file_path:
            # Trên Android đôi khi path là None, phải copy
            append_log(f"Không lấy được đường dẫn file. Tên file: {e.files[0].name}")
            return

        try:
            append_log(f"\n{'='*40}\nĐã chọn: {file_path}\n{'='*40}")
            output = process_excel_file(file_path, append_log)
            page.snack_bar = ft.SnackBar(ft.Text(f"Xong! Đã tạo {os.path.basename(output)}"))
            page.snack_bar.open = True
            page.update()
        except Exception as ex:
            append_log(f"\n!!! LỖI: {ex}")
            import traceback
            append_log(traceback.format_exc())
            page.snack_bar = ft.SnackBar(ft.Text(f"Lỗi: {ex}"), bgcolor="red")
            page.snack_bar.open = True
            page.update()

    file_picker = ft.FilePicker(on_result=on_pick_result)
    page.overlay.append(file_picker)

    btn_pick = ft.ElevatedButton(
        "1. Chọn File Excel",
        icon=ft.Icons.FILE_OPEN,
        on_click=lambda _: file_picker.pick_files(
            allowed_extensions=["xlsx"],
            allow_multiple=False
        ),
        style=ft.ButtonStyle(bgcolor="#2196F3", color="white")
    )

    btn_clear = ft.TextButton("Xóa log", on_click=lambda _: [setattr(log_text, 'value', ''), page.update()])

    page.add(
        ft.Text("TOOL TÁCH SHEET - BẢN FIX FLET", size=20, weight="bold"),
        ft.Text("Bản này fix lỗi EOF when reading a line trên điện thoại", size=12, color="grey"),
        ft.Divider(),
        ft.Row([btn_pick, btn_clear]),
        log_container
    )

ft.app(target=main)
